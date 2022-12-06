# -*- coding: utf-8 -*-
"""
  Created on Wednesday November 30, 2022
  @author: Fedor Alexander Steeman, NHMD
  Copyright 2022 Natural History Museum of Denmark (NHMD)
  Licensed under the Apache License, Version 2.0 (the "License");
  you may not use this file except in compliance with the License.
  You may obtain a copy of the License at
  http://www.apache.org/licenses/LICENSE-2.0
  Unless required by applicable law or agreed to in writing, software distributed under the License is distributed on an "AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied. See the License for the specific language governing permissions and limitations under the License.

  PURPOSE: Parse excel file containing locality records and using lat/long coordinaties to determine Danish municipality and adding that information to the output file. 
"""

import time
import pandas as pd
import openpyxl as px 
import tempfile as tf
from datetime import datetime as dt
import dawa_interface

class KommuneFinder():
  """
  Fetches Danish municipality name from DAGI webservice API on the basis of latitude and longitude coordinates. 
  """

  def __init__(self) -> None:
    """
    Constructor initializing DAGI web API wrapper. 
    """
    self.di = dawa_interface.DawaInterface()

  def findKommuner(self, filename, number):
    """
    Processes excel file containing locality records and matching Danish municipality name using latitude and longitude. 
    Generates output file with same records but adding "kommune" column containing municipality name. 
    TODO The below code is expecting the latitude longitude to be in fixed columns. 
    CONTRACT 
      filename (String) : filename of the file to be processed. 
      number (int) : Serial number added to the filename 
    """
    if number < 10: 
      strNumber = f'0{number}'
    else:
      strNumber = f'{number}'
    outputpath = self.generateFilename(f'danskestednavne{strNumber}', 'xlsx', 'files')
    wbookOut = px.Workbook()
    sheetOut = wbookOut.active

    fullfilename = f'{filename}_{number}.xlsx'

    print(f' - Reading in Excel file {fullfilename} using openpyxl...')
    start = time.time()
    wbook = px.load_workbook(f'files\{fullfilename}', read_only=True)
    sheetIn = wbook.active
    end = time.time()
    timeElapsed = end - start
    print(f'Time elapsed: {timeElapsed} ')

    # Generate output columns
    headerRow = sheetIn[1]
    colindex = 1
    for cell in headerRow:
    #for col in sheetIn.iter_cols():
      sheetOut.insert_cols(idx=colindex)
      colindex = colindex + 1 
    sheetOut.insert_cols(idx=colindex)

    # Insert headers
    columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']
    for c in columns:
      clin = sheetIn[f'{c}1']
      clout = sheetOut[f'{c}1']
      clout.value = clin.value
    cl = sheetOut['J1']
    cl.value='kommune'

    print(f' - Processing file...')
    start = time.time()
    # iterate through input sheet and set values into output sheet
    rowindex = 2
    for row in sheetIn.iter_rows(min_row=2, min_col=1, max_row=120000, max_col=9):
      colindex = 0
      sheetOut._add_row()
      for cell in row:
        colLetter = columns[colindex]
        cx = f'{colLetter}{rowindex}'
        #print(f'{cx}:{cell.value}', end="|")
        print('.', end = '')
        sheetOut[cx].value = cell.value
        colindex = colindex + 1      
      kommune = self.di.getKommuneName(sheetOut[f'C{rowindex}'].value, sheetOut[f'B{rowindex}'].value)
      if kommune != '': print('*', end = '')
      else: print('!', end = '')
      #print(f'Found kommune: {kommune}')
      sheetOut[f'J{rowindex}'].value = kommune 
      rowindex = rowindex + 1 
      #print()
    print()
    end = time.time()
    timeElapsed = end - start
    print(f'Time elapsed: {timeElapsed} ')
    print(f'Saving file to: {outputpath}')
    wbookOut.save(outputpath)
    wbookOut.close()

  def generateFilename(self, object_name, file_type, file_path):
    # Generic method for generating random filename including path denoting object type and timestamp 
    # CONTRACT 
    #    object_name (String): Name of the object i.e. table to be generated file name for 
    #    file_type (String): Extension of the file type to be exported to 
    #    file_path (String): Path to the folder to be exported to 
    #    RETURNS Text string denoting the path and filename as specified. 
    return tf.NamedTemporaryFile(prefix='%s-export_%s'%(object_name,dt.now().strftime("%Y%m%d%H%M_")), suffix='.%s'%file_type, dir=file_path).name

kf = KommuneFinder()

kf.findKommuner('DanskeStednavne-Processed', 5)
